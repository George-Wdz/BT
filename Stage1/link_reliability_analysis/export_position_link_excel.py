#!/usr/bin/env python3
"""Export position-link analysis artifacts to one reviewable workbook."""
from __future__ import annotations

import argparse
import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


FILES = {
    "geometry_summary": "geometry_dropout_summary.csv",
    "monthly_summary": "position_link_monthly_summary.csv",
    "satellite_summary": "satellite_position_link_summary.csv",
    "phy_quality": "phy_quality_summary.csv",
    "position_quality": "position_quality_summary.csv",
    "session_details": "position_link_passes.csv",
    "visibility_opportunities": "position_visibility_opportunities.csv",
}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--artifact-dir", type=Path, required=True)
    parser.add_argument("--output-path", type=Path, required=True)
    args = parser.parse_args()
    args.output_path.parent.mkdir(parents=True, exist_ok=True)

    dashboard = json.loads((args.artifact_dir / "position_link_dashboard.json").read_text(encoding="utf-8"))
    overview = pd.DataFrame([
        {"指标": key, "数值": value} for key, value in dashboard["overview"].items()
    ])
    method = pd.DataFrame([
        {"项目": key, "说明": value} for key, value in dashboard["method"].items()
    ])
    with pd.ExcelWriter(args.output_path, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="overview", index=False)
        method.to_excel(writer, sheet_name="method", index=False)
        for sheet, filename in FILES.items():
            pd.read_csv(args.artifact_dir / filename).to_excel(
                writer, sheet_name=sheet[:31], index=False
            )

    from openpyxl import load_workbook
    workbook = load_workbook(args.output_path)
    fill, font = PatternFill("solid", fgColor="1769AA"), Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill, cell.font = fill, font
            cell.alignment = Alignment(horizontal="center")
        for cells in sheet.columns:
            width = min(max(max((len(str(cell.value or "")) for cell in cells[:1500]), default=8) + 2, 11), 32)
            sheet.column_dimensions[cells[0].column_letter].width = width
    workbook.save(args.output_path)
    print(args.output_path)


if __name__ == "__main__":
    main()
