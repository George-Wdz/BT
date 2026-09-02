#!/usr/bin/env python3
"""Export traceable raw PHY rows into one Excel workbook per month."""
from __future__ import annotations

import argparse
import hashlib
import json
import sqlite3
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from analyze_raw_link_reliability import classify_phy, db_time_bound, sha256


MAX_DATA_ROWS = 900_000


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def write_manifest(output_dir: Path) -> None:
    artifact_dir = output_dir.parent
    monthly = pd.read_csv(artifact_dir / "monthly_summary.csv").set_index("month")
    files = []
    for path in sorted(output_dir.glob("raw_phy_*.xlsx")):
        month = path.stem.removeprefix("raw_phy_")
        files.append({
            "month": month,
            "path": str(path),
            "bytes": path.stat().st_size,
            "sha256": file_sha256(path),
            "expected_raw_rows": int(monthly.loc[month, "raw_phy_rows"]),
        })
    manifest = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "note": "Each workbook contains provenance plus all traceable raw PHY rows for the month; April uses multiple sheets because of Excel's row limit.",
        "files": files,
    }
    path = output_dir / "monthly_phy_excel_manifest.json"
    path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"saved={path}")


class MonthlyWriter:
    def __init__(self, output_dir: Path, month: str, provenance: list[dict]):
        self.path = output_dir / f"raw_phy_{month}.xlsx"
        self.book = Workbook(write_only=True)
        meta = self.book.create_sheet("provenance")
        meta.append(["field", "value"])
        meta.append(["export_month", month])
        meta.append(["generated_at", datetime.now().isoformat(timespec="seconds")])
        for index, item in enumerate(provenance):
            meta.append([f"source_{index}", json.dumps(item, ensure_ascii=False)])
        self.sheet_index = 0
        self.rows = 0
        self.sheet = None
        self.header = None

    def append_frame(self, frame: pd.DataFrame):
        if self.header is None:
            self.header = list(frame.columns)
        for row in frame.itertuples(index=False, name=None):
            if self.sheet is None or self.rows >= MAX_DATA_ROWS:
                self.sheet_index += 1
                self.sheet = self.book.create_sheet(f"phy_{self.sheet_index:02d}")
                self.sheet.append(self.header)
                self.rows = 0
            self.sheet.append(row)
            self.rows += 1

    def close(self):
        self.book.save(self.path)
        print(f"saved={self.path}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--early-db", type=Path, default=Path("/home/wdz/BT/db_backups/satellite_data_20260527_100641_before_clean.db"))
    parser.add_argument("--raw-phy-csv", type=Path, default=Path("/home/wdz/satellite_data/工控机采集的原始备份数据/phy_data.csv"))
    parser.add_argument("--output-dir", type=Path, default=Path(__file__).resolve().parent / "artifacts" / "monthly_phy_excel")
    parser.add_argument("--chunksize", type=int, default=100_000)
    parser.add_argument("--manifest-only", action="store_true")
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    if args.manifest_only:
        write_manifest(args.output_dir)
        return
    cutover = db_time_bound(args.early_db, "phy_data", "localTime")
    provenance = [
        {"path": str(args.early_db), "sha256": sha256(args.early_db), "selection": f"localTime <= {cutover.isoformat()}"},
        {"path": str(args.raw_phy_csv), "sha256": sha256(args.raw_phy_csv), "selection": f"localTime >= {cutover.isoformat()}"},
        {"cleaning_rules": "/home/wdz/satellite_data/server.py", "note": "raw fields retained; clean_reason applies field rules only and is derived non-destructively; global 1 s near-duplicates are reported in the analysis artifacts"},
    ]

    def chunks():
        with sqlite3.connect(f"file:{args.early_db}?mode=ro", uri=True) as conn:
            query = "SELECT * FROM phy_data WHERE datetime(localTime) <= datetime(?) ORDER BY localTime,id"
            for frame in pd.read_sql_query(query, conn, params=[cutover.isoformat()], chunksize=args.chunksize):
                frame["source_file"] = str(args.early_db); frame["source_record_id"] = frame["id"]
                yield frame
        for frame in pd.read_csv(args.raw_phy_csv, chunksize=args.chunksize, low_memory=False):
            timestamp = pd.to_datetime(frame.localTime, errors="coerce", format="mixed")
            frame = frame.loc[timestamp > cutover].copy()
            if frame.empty:
                continue
            frame["source_file"] = str(args.raw_phy_csv); frame["source_record_id"] = frame["id"]
            yield frame

    writers = {}
    for frame in chunks():
        frame["clean_reason"] = classify_phy(frame)
        frame["clean_status"] = frame.clean_reason.map(lambda value: "valid" if value == "valid" else ("no_satellite_lock" if value == "no_satellite_lock" else "invalid"))
        frame["month"] = pd.to_datetime(frame.localTime, errors="coerce", format="mixed").dt.strftime("%Y-%m")
        for month, part in frame.groupby("month", sort=True):
            if month not in writers:
                writers[month] = MonthlyWriter(args.output_dir, month, provenance)
            writers[month].append_frame(part.drop(columns="month"))
    for writer in writers.values():
        writer.close()
    write_manifest(args.output_dir)


if __name__ == "__main__":
    main()
