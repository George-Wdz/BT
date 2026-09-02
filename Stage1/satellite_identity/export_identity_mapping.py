#!/usr/bin/env python3
"""Export a concise receiver ephemeris ID to physical NORAD ID mapping."""
from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


def style_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1769AA")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        widths = {"A": 14, "B": 18, "C": 20, "D": 20, "E": 14}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--mapping-csv", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--latest-let-version", default="0727")
    args = parser.parse_args()

    source = Path(args.mapping_csv).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    frame = pd.read_csv(source, dtype={"let_version": str})
    frame = frame.loc[frame["let_version"].eq(str(args.latest_let_version))].copy()
    if frame.empty:
        raise ValueError(f"no rows for latest LET version {args.latest_let_version}")
    frame = frame.sort_values(["let_version", "raw_satellite_id"]).drop_duplicates(
        ["let_version", "raw_satellite_id"], keep="first"
    )
    mapped = frame.loc[frame["status"].eq("accepted")].copy()
    pending = frame.loc[~frame["status"].eq("accepted")].copy()

    def concise(rows: pd.DataFrame, include_status: bool = False) -> pd.DataFrame:
        result = rows[
            ["let_version", "raw_satellite_id", "norad_id", "physical_name"]
        ].copy()
        result.columns = [
            "LET版本", "终端星历ID", "卫星物理ID_NORAD", "千帆卫星编号"
        ]
        result["终端星历ID"] = result["终端星历ID"].astype("Int64")
        result["卫星物理ID_NORAD"] = result["卫星物理ID_NORAD"].astype("Int64")
        if include_status:
            result["映射状态"] = rows["status"].map(
                {"accepted": "已映射", "provisional": "待确认", "unresolved": "待映射"}
            ).to_numpy()
        return result

    mapped_output = concise(mapped)
    pending_output = concise(pending, include_status=True)
    pending_output["卫星物理ID_NORAD"] = pd.Series(
        pd.NA, index=pending_output.index, dtype="Int64"
    )
    pending_output["千帆卫星编号"] = pd.NA
    all_output = concise(frame, include_status=True)
    unresolved = all_output["映射状态"].ne("已映射")
    all_output.loc[unresolved, ["卫星物理ID_NORAD", "千帆卫星编号"]] = pd.NA

    excel_path = output_dir / "terminal_ephemeris_to_physical_id_mapping.xlsx"
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        mapped_output.to_excel(writer, sheet_name="已映射", index=False)
        pending_output.to_excel(writer, sheet_name="待映射", index=False)
        all_output.to_excel(writer, sheet_name="全部", index=False)
    style_workbook(excel_path)

    sqlite_path = output_dir / "terminal_ephemeris_to_physical_id_mapping.sqlite3"
    if sqlite_path.exists():
        sqlite_path.unlink()
    with sqlite3.connect(sqlite_path) as connection:
        mapped_output.rename(columns={
            "LET版本": "let_version",
            "终端星历ID": "terminal_ephemeris_id",
            "卫星物理ID_NORAD": "physical_norad_id",
            "千帆卫星编号": "physical_name",
        }).to_sql("mapped_satellites", connection, index=False)
        pending_output.rename(columns={
            "LET版本": "let_version",
            "终端星历ID": "terminal_ephemeris_id",
            "卫星物理ID_NORAD": "physical_norad_id",
            "千帆卫星编号": "physical_name",
            "映射状态": "mapping_status",
        }).to_sql("pending_satellites", connection, index=False)
        all_output.rename(columns={
            "LET版本": "let_version",
            "终端星历ID": "terminal_ephemeris_id",
            "卫星物理ID_NORAD": "physical_norad_id",
            "千帆卫星编号": "physical_name",
            "映射状态": "mapping_status",
        }).to_sql("all_satellites", connection, index=False)
        connection.executescript(
            """
            CREATE UNIQUE INDEX idx_mapped_version_id
              ON mapped_satellites(let_version, terminal_ephemeris_id);
            CREATE UNIQUE INDEX idx_pending_version_id
              ON pending_satellites(let_version, terminal_ephemeris_id);
            CREATE UNIQUE INDEX idx_all_version_id
              ON all_satellites(let_version, terminal_ephemeris_id);
            """
        )

    print(f"source={source}")
    print(f"mapped={len(mapped_output)} pending={len(pending_output)} total={len(all_output)}")
    print(f"excel={excel_path}")
    print(f"sqlite={sqlite_path}")


if __name__ == "__main__":
    main()
