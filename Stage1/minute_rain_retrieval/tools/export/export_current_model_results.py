#!/usr/bin/env python3
"""Export current minute-rain model evaluation and typhoon results."""
from __future__ import annotations

import argparse
import json
import sqlite3
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


RAIN_THRESHOLD = 0.005
PROBABILITY_THRESHOLD = 0.5
TYPHOONS = {
    "typhoon_jul": ("2026-07-11 00:00:00", "2026-07-13 00:00:00"),
    "typhoon_aug": ("2026-08-09 00:00:00", "2026-08-11 00:00:00"),
}
DETAIL_COLUMNS = [
    "时间", "数据划分", "终端ID", "卫星ID", "采样点数", "卫星数量",
    "降雨概率", "真实分钟降雨_mm", "反演分钟降雨_mm", "绝对误差_mm", "推理模式",
]


def sample_metadata(dataset_path: Path) -> pd.DataFrame:
    archive = np.load(dataset_path, allow_pickle=True)
    samples = archive["samples"].tolist()
    splits = archive["splits"].astype(str)
    rows = []
    for sample, split in zip(samples, splits):
        satellites = np.asarray(sample["satellite_ids"], dtype=np.int64)
        valid = satellites[satellites >= 0]
        if len(valid):
            values, counts = np.unique(valid, return_counts=True)
            dominant = int(values[np.argmax(counts)])
        else:
            dominant = None
        rows.append({
            "anchor_ns": int(sample["anchor_time_ns"]),
            "数据划分": str(split),
            "卫星ID": dominant,
            "采样点数": int(sample["point_count"]),
            "卫星数量": int(sample["satellite_count"]),
        })
    return pd.DataFrame(rows).drop_duplicates("anchor_ns", keep="last")


def load_evaluation(output_dir: Path, metadata: pd.DataFrame) -> dict[str, pd.DataFrame]:
    frames = {}
    for split in ("val", "test"):
        frame = pd.read_csv(output_dir / f"{split}_predictions.csv")
        frame["时间"] = pd.to_datetime(frame.pop("anchor_time"), errors="coerce")
        frame["anchor_ns"] = frame["时间"].astype("int64")
        frame = frame.merge(metadata, on="anchor_ns", how="left", validate="one_to_one")
        frame["数据划分"] = split
        frame["终端ID"] = "01-31-0005-0001"
        frame["降雨概率"] = frame.pop("rain_probability")
        frame["真实分钟降雨_mm"] = frame.pop("true_minute_rainfall_mm").round(2)
        frame["反演分钟降雨_mm"] = frame.pop("pred_minute_rainfall_mm")
        frame["绝对误差_mm"] = (
            frame["反演分钟降雨_mm"] - frame["真实分钟降雨_mm"]
        ).abs()
        frame["推理模式"] = "完整位置模型"
        frames[split] = frame[DETAIL_COLUMNS].sort_values("时间")
    return frames


def load_history(history_db: Path, primary_meta: pd.DataFrame,
                 fallback_meta: pd.DataFrame) -> pd.DataFrame:
    query = """
        WITH ranked AS (
          SELECT *, ROW_NUMBER() OVER (
            PARTITION BY terminal_id, pass_end ORDER BY datetime(inferred_at) DESC, id DESC
          ) AS row_rank
          FROM rain_retrieval_passes
          WHERE observed_available=1 AND observed_rainfall_mm IS NOT NULL
        )
        SELECT terminal_id, satellite_id, pass_end, points, rain_probability,
               reported_rainfall_mm, observed_rainfall_mm, transfer_mode
        FROM ranked WHERE row_rank=1 ORDER BY datetime(pass_end), terminal_id
    """
    with sqlite3.connect(f"file:{history_db}?mode=ro", uri=True) as connection:
        frame = pd.read_sql_query(query, connection)
    frame["时间"] = pd.to_datetime(frame.pop("pass_end"), errors="coerce")
    frame["anchor_ns"] = frame["时间"].astype("int64")
    frame["推理模式"] = np.where(
        frame["transfer_mode"].fillna("").str.contains("fallback_no_position"),
        "无位置回退", "完整位置模型",
    )
    primary_split = primary_meta.set_index("anchor_ns")["数据划分"].to_dict()
    fallback_split = fallback_meta.set_index("anchor_ns")["数据划分"].to_dict()
    frame["数据划分"] = [
        (fallback_split if mode == "无位置回退" else primary_split).get(int(anchor), "不在训练NPZ")
        for anchor, mode in zip(frame["anchor_ns"], frame["推理模式"])
    ]
    frame = frame.rename(columns={
        "terminal_id": "终端ID", "satellite_id": "卫星ID", "points": "采样点数",
        "rain_probability": "降雨概率", "reported_rainfall_mm": "反演分钟降雨_mm",
        "observed_rainfall_mm": "真实分钟降雨_mm",
    })
    frame["卫星数量"] = np.nan
    frame["真实分钟降雨_mm"] = frame["真实分钟降雨_mm"].round(2)
    frame["绝对误差_mm"] = (
        frame["反演分钟降雨_mm"] - frame["真实分钟降雨_mm"]
    ).abs()
    return frame[DETAIL_COLUMNS].dropna(subset=["时间"]).sort_values(["时间", "终端ID"])


def metrics(name: str, frame: pd.DataFrame) -> dict:
    true = frame["真实分钟降雨_mm"].to_numpy(float)
    prediction = frame["反演分钟降雨_mm"].to_numpy(float)
    probability = frame["降雨概率"].to_numpy(float)
    rainy = true > RAIN_THRESHOLD
    predicted_rain = probability >= PROBABILITY_THRESHOLD
    tp = int((rainy & predicted_rain).sum())
    fp = int((~rainy & predicted_rain).sum())
    fn = int((rainy & ~predicted_rain).sum())
    tn = int((~rainy & ~predicted_rain).sum())
    absolute = np.abs(prediction - true)
    return {
        "数据表": name,
        "样本数": len(frame),
        "真实有雨样本": int(rainy.sum()),
        "MAE_mm": float(absolute.mean()) if len(frame) else np.nan,
        "有雨MAE_mm": float(absolute[rainy].mean()) if rainy.any() else np.nan,
        "分类准确率": (tp + tn) / max(len(frame), 1),
        "精确率": tp / max(tp + fp, 1),
        "召回率": tp / max(tp + fn, 1),
        "F1": 2 * tp / max(2 * tp + fp + fn, 1),
        "假警率": fp / max(fp + tn, 1),
        "TP": tp, "FP": fp, "FN": fn, "TN": tn,
    }


def style(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    fill = PatternFill("solid", fgColor="1769AA")
    font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill, cell.font = fill, font
            cell.alignment = Alignment(horizontal="center")
        for cells in sheet.columns:
            width = min(max(max((len(str(c.value or "")) for c in cells[:2000]), default=8) + 2, 11), 27)
            sheet.column_dimensions[cells[0].column_letter].width = width
        for row in sheet.iter_rows(min_row=2):
            if row and hasattr(row[0].value, "year"):
                row[0].number_format = "yyyy-mm-dd hh:mm:ss"
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dataset-path", required=True)
    parser.add_argument("--fallback-dataset-path", required=True)
    parser.add_argument("--model-output-dir", required=True)
    parser.add_argument("--history-db", required=True)
    parser.add_argument("--output-path", required=True)
    args = parser.parse_args()

    dataset = Path(args.dataset_path).resolve()
    fallback_dataset = Path(args.fallback_dataset_path).resolve()
    model_output = Path(args.model_output_dir).resolve()
    history_db = Path(args.history_db).resolve()
    output = Path(args.output_path).resolve()
    output.parent.mkdir(parents=True, exist_ok=True)

    primary_meta = sample_metadata(dataset)
    fallback_meta = sample_metadata(fallback_dataset)
    evaluation = load_evaluation(model_output, primary_meta)
    val_test = pd.concat([evaluation["val"], evaluation["test"]]).sort_values("时间")
    history = load_history(history_db, primary_meta, fallback_meta)

    tables = {
        "val_all": evaluation["val"],
        "test_all": evaluation["test"],
        "val_test_all": val_test,
        "val_rainy": evaluation["val"].loc[evaluation["val"]["真实分钟降雨_mm"] > RAIN_THRESHOLD],
        "test_rainy": evaluation["test"].loc[evaluation["test"]["真实分钟降雨_mm"] > RAIN_THRESHOLD],
        "val_test_rainy": val_test.loc[val_test["真实分钟降雨_mm"] > RAIN_THRESHOLD],
    }
    for name, (start, end) in TYPHOONS.items():
        period = history.loc[history["时间"].ge(start) & history["时间"].lt(end)].copy()
        tables[f"{name}_all"] = period
        tables[f"{name}_rainy"] = period.loc[period["真实分钟降雨_mm"] > RAIN_THRESHOLD]

    summary = pd.DataFrame([metrics(name, frame) for name, frame in tables.items()])
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="metrics", index=False)
        for name, frame in tables.items():
            export = frame.copy()
            for column in ("降雨概率", "反演分钟降雨_mm", "绝对误差_mm"):
                export[column] = export[column].round(6)
            export.to_excel(writer, sheet_name=name[:31], index=False)
    style(output)

    summary_path = output.with_suffix(".metrics.json")
    summary_path.write_text(
        json.dumps(summary.to_dict("records"), ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps({
        "workbook": str(output), "metrics": str(summary_path),
        "tables": {name: len(frame) for name, frame in tables.items()},
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
