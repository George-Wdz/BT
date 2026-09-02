#!/usr/bin/env python3
"""Export minute-rain history and two typhoon periods for audit."""
from __future__ import annotations

import argparse
import json
import sqlite3
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


PERIODS = {
    "aug09_aug11_all": ("2026-08-09 00:00:00", "2026-08-11 00:00:00"),
    "jul11_jul13_all": ("2026-07-11 00:00:00", "2026-07-13 00:00:00"),
}

COLUMNS = {
    "pass_start": "分钟窗口开始",
    "pass_end": "分钟锚点时间",
    "terminal_id": "终端ID",
    "satellite_id": "主卫星ID",
    "points": "PHY点数",
    "rain_probability": "反演降雨概率",
    "reported_rainfall_mm": "反演分钟降雨量_mm",
    "observed_rainfall_mm": "真实分钟降雨量_mm",
    "absolute_error_mm": "绝对误差_mm",
    "pred_rainfall_mm": "原始模型输出_mm",
    "inference_mode": "推理模式",
    "checkpoint_split": "训练数据划分",
    "used_for_training": "是否参与5比1模型训练",
}


def load_history(path: Path) -> pd.DataFrame:
    query = """
        WITH ranked AS (
            SELECT *,
                   ROW_NUMBER() OVER (
                       PARTITION BY terminal_id, pass_end
                       ORDER BY datetime(inferred_at) DESC, id DESC
                   ) AS row_rank
            FROM rain_retrieval_passes
            WHERE observed_available = 1
              AND observed_rainfall_mm IS NOT NULL
        )
        SELECT terminal_id, satellite_id, pass_start, pass_end, points,
               rain_probability, pred_rainfall_mm, reported_rainfall_mm,
               observed_rainfall_mm, transfer_mode, position_source, inferred_at
        FROM ranked
        WHERE row_rank = 1
        ORDER BY datetime(pass_end), terminal_id
    """
    with sqlite3.connect(f"file:{path}?mode=ro", uri=True) as connection:
        frame = pd.read_sql_query(query, connection)
    frame["pass_start"] = pd.to_datetime(frame["pass_start"], errors="coerce")
    frame["pass_end"] = pd.to_datetime(frame["pass_end"], errors="coerce")
    for column in (
        "rain_probability",
        "pred_rainfall_mm",
        "reported_rainfall_mm",
        "observed_rainfall_mm",
    ):
        frame[column] = pd.to_numeric(frame[column], errors="coerce")
    frame["absolute_error_mm"] = (
        frame["reported_rainfall_mm"] - frame["observed_rainfall_mm"]
    ).abs()
    frame["inference_mode"] = np.where(
        frame["transfer_mode"].fillna("").str.contains("fallback_no_position"),
        "无位置回退", "完整位置模型",
    )
    return frame.dropna(subset=["pass_start", "pass_end"])


def load_checkpoint_splits(path: Path) -> dict[int, str]:
    archive = np.load(path, allow_pickle=True)
    samples = archive["samples"].tolist()
    splits = archive["splits"].astype(str)
    return {
        int(sample["anchor_time_ns"]): str(split)
        for sample, split in zip(samples, splits)
    }


def attach_split(
    frame: pd.DataFrame,
    split_by_time: dict[int, str],
    fallback_split_by_time: dict[int, str],
) -> pd.DataFrame:
    result = frame.copy()
    anchor_ns = result["pass_end"].astype("int64")
    result["checkpoint_split"] = [
        (
            fallback_split_by_time if mode == "无位置回退" else split_by_time
        ).get(int(timestamp), "not_in_training_npz")
        for timestamp, mode in zip(anchor_ns, result["inference_mode"])
    ]
    result["used_for_training"] = (
        result["terminal_id"].eq("01-31-0005-0001")
        & result["checkpoint_split"].eq("train")
    )
    return result


def select_period(frame: pd.DataFrame, start: str, end: str) -> pd.DataFrame:
    return frame.loc[
        frame["pass_end"].ge(pd.Timestamp(start))
        & frame["pass_end"].lt(pd.Timestamp(end))
    ].copy()


def display_frame(frame: pd.DataFrame) -> pd.DataFrame:
    output = frame[list(COLUMNS)].rename(columns=COLUMNS).copy()
    output["反演降雨概率"] = output["反演降雨概率"].round(6)
    for column in (
        "反演分钟降雨量_mm",
        "绝对误差_mm",
        "原始模型输出_mm",
    ):
        output[column] = output[column].round(6)
    output["真实分钟降雨量_mm"] = output["真实分钟降雨量_mm"].round(2)
    output["是否参与5比1模型训练"] = output["是否参与5比1模型训练"].map(
        {True: "是", False: "否"}
    )
    return output


def summarize(name: str, frame: pd.DataFrame) -> dict:
    true = frame["observed_rainfall_mm"].to_numpy(dtype=float)
    prediction = frame["reported_rainfall_mm"].to_numpy(dtype=float)
    return {
        "table": name,
        "records": int(len(frame)),
        "terminals": int(frame["terminal_id"].nunique()),
        "true_rainy_records": int((true > 0).sum()),
        "predicted_rainy_records": int((prediction > 0).sum()),
        "true_rainfall_sum_mm": round(float(true.sum()), 6),
        "predicted_rainfall_sum_mm": round(float(prediction.sum()), 6),
        "mae_mm": round(float(np.abs(prediction - true).mean()), 6) if len(frame) else None,
        "training_rows": int(frame["used_for_training"].sum()),
        "validation_rows": int(frame["checkpoint_split"].eq("val").sum()),
        "test_rows": int(frame["checkpoint_split"].eq("test").sum()),
        "not_in_training_npz": int(
            frame["checkpoint_split"].eq("not_in_training_npz").sum()
        ),
    }


def style_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F6D4F")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in sheet.columns:
            values = [str(cell.value or "") for cell in column_cells[:2000]]
            width = min(max(max(map(len, values), default=8) + 2, 10), 34)
            sheet.column_dimensions[column_cells[0].column_letter].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column in (1, 2) and isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"
                elif isinstance(cell.value, float):
                    cell.number_format = "0.000000"
    workbook.save(path)


def write_report(
    path: Path,
    summaries: list[dict],
    history_path: Path,
    dataset_path: Path,
    workbook_path: Path,
) -> None:
    header = [
        "# 分钟降雨反演与台风时段导出",
        "",
        f"- 生成时间：{datetime.now().isoformat(timespec='seconds')}",
        f"- 分钟历史库：`{history_path}`",
        f"- 训练数据集：`{dataset_path}`",
        f"- Excel：`{workbook_path}`",
        "- `主卫星ID` 为该分钟 PHY 点数最多的卫星；模型实际输入可能包含多颗卫星。",
        "- `反演分钟降雨量_mm` 为服务经过降雨概率阈值后的最终输出；`原始模型输出_mm` 为阈值前输出。",
        "- `是否参与5比1模型训练=是` 仅表示终端 001 且该分钟属于 checkpoint 的 train split。",
        "",
        "## 汇总",
        "",
        "| 表格 | 记录数 | 真实有雨 | 模型有雨 | 真实累计/mm | 反演累计/mm | MAE/mm | 训练行 | 验证行 | 测试行 | 不在训练NPZ |",
        "|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for item in summaries:
        header.append(
            "| {table} | {records} | {true_rainy_records} | "
            "{predicted_rainy_records} | {true_rainfall_sum_mm:.6f} | "
            "{predicted_rainfall_sum_mm:.6f} | {mae} | {training_rows} | "
            "{validation_rows} | {test_rows} | {not_in_training_npz} |".format(
                **item,
                mae=(f"{item['mae_mm']:.6f}" if item["mae_mm"] is not None else "-"),
            )
        )
    path.write_text("\n".join(header) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--history-db", required=True)
    parser.add_argument("--dataset-path", required=True)
    parser.add_argument("--fallback-dataset-path", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument(
        "--output-name", default="minute_rainfall_typhoon_comparison_20260819.xlsx"
    )
    args = parser.parse_args()

    history_path = Path(args.history_db).expanduser().resolve()
    dataset_path = Path(args.dataset_path).expanduser().resolve()
    fallback_dataset_path = Path(args.fallback_dataset_path).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    history = attach_split(
        load_history(history_path), load_checkpoint_splits(dataset_path),
        load_checkpoint_splits(fallback_dataset_path),
    )
    frames = {
        "all_rainy": history.loc[history["observed_rainfall_mm"] > 0].copy(),
        **{
            name: select_period(history, start, end)
            for name, (start, end) in PERIODS.items()
        },
    }
    frames["aug09_aug11_rainy"] = frames["aug09_aug11_all"].loc[
        frames["aug09_aug11_all"]["observed_rainfall_mm"] > 0
    ].copy()
    frames["jul11_jul13_rainy"] = frames["jul11_jul13_all"].loc[
        frames["jul11_jul13_all"]["observed_rainfall_mm"] > 0
    ].copy()

    workbook_path = output_dir / args.output_name
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        for name, frame in frames.items():
            display_frame(frame).to_excel(writer, sheet_name=name[:31], index=False)
        summary_frame = pd.DataFrame(
            [summarize(name, frame) for name, frame in frames.items()]
        )
        summary_frame.to_excel(writer, sheet_name="summary", index=False)
    style_workbook(workbook_path)

    for name in ("all_rainy", "aug09_aug11_all", "jul11_jul13_all"):
        display_frame(frames[name]).to_csv(
            output_dir / f"{name}.csv", index=False, encoding="utf-8-sig"
        )

    summaries = [summarize(name, frame) for name, frame in frames.items()]
    report_path = output_dir / "minute_rainfall_typhoon_comparison.md"
    write_report(
        report_path, summaries, history_path, dataset_path, workbook_path
    )
    (output_dir / "summary.json").write_text(
        json.dumps(summaries, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps({
        "workbook": str(workbook_path),
        "report": str(report_path),
        "tables": {name: len(frame) for name, frame in frames.items()},
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
